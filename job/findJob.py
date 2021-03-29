# -*- coding: utf-8 -*-

import json
import time

import numpy as np
from pyecharts import options as opts
from pyecharts.charts import BMap
from pyecharts.globals import ChartType, SymbolType, ThemeType
import requests

from selenium import webdriver
from bs4 import BeautifulSoup
from lxml import etree
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

from openpyxl import Workbook

# 解决中文显示不出来
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']

salary_range = {
    '3k以下': {'min': 0, 'max': 3000, 'count': 0},
    '3k-5k': {'min': 3000, 'max': 5000, 'count': 0},
    '5k-6k': {'min': 5000, 'max': 6000, 'count': 0},
    '6k-8k': {'min': 6000, 'max': 8000, 'count': 0},
    '8k-10k': {'min': 8000, 'max': 10000, 'count': 0},
    '10k-15k': {'min': 10000, 'max': 15000, 'count': 0},
    '15k-20k': {'min': 15000, 'max': 20000, 'count': 0},
    '20k-30k': {'min': 20000, 'max': 30000, 'count': 0},
    '30k以上': {'min': 30000, 'max': 300000, 'count': 0},
}

def createExcel():
    # 新建工作簿
    workbook = Workbook()
    sheet1 = workbook.active
    # 给新建的工作表改名
    sheet1.title = '招聘岗位薪资待遇'
    # workbook = xlwt.Workbook(encoding='utf-8')  # 新建工作簿
    # sheet1 = workbook.add_sheet("招聘岗位薪资待遇")  # 新建sheet

    # sheet1.write(0, 0, "公司")  # 第1行第1列数据
    sheet1.cell(row=1, column=1, value="职位")
    sheet1.cell(row=1, column=2, value="职务类别1")
    sheet1.cell(row=1, column=3, value="职务类别2")
    sheet1.cell(row=1, column=4, value="职务类别3")
    sheet1.cell(row=1, column=5, value="薪资")
    sheet1.cell(row=1, column=6, value="年限要求")
    sheet1.cell(row=1, column=7, value="学历要求")
    sheet1.cell(row=1, column=8, value="区域")
    # sheet1.cell(row=1, column=3, value="所属区域")
    # sheet1.cell(row=1, column=3, value="公司规模")

    # 读取json文件
    # 先按公司将相同职位合并

    jobs = getJobs()

    count = 2
    for data in jobs:
        # print(data)
        sheet1.cell(row=count, column=1, value=data['positionName'])  # "职位"
        sheet1.cell(row=count, column=2, value=data['firstType'])  # "职务类别1")
        sheet1.cell(row=count, column=3, value=data['secondType'])  # "职务类别2")
        sheet1.cell(row=count, column=4, value=data['thirdType'])  # "职务类别3")
        sheet1.cell(row=count, column=5, value=data['salary'])  # "薪资")
        sheet1.cell(row=count, column=6, value=data['workYear'])  # "年限要求")
        sheet1.cell(row=count, column=7, value=data['education'])  # "学历要求")
        # sheet1.cell(row=1, column=8, value=data['skillLables'])  # "技能要求")
        sheet1.cell(row=count, column=8, value=data['district'])  # "区域")

        count +=1

    # sheet1.write(0, 1, "学号")  # 第1行第2列数据
    # sheet1.write(1, 0, "张三")  # 第2行第1列数据
    # sheet1.write(1, 1, "036")  # 第2行第2列数据
    workbook.save(r'jobs.xlsx')  # 保存


def getJobs():
    f = open("original_jobs_data.txt", "r")
    jobs = json.loads(f.read())
    newJobs = {}
    for job in jobs:
        newJobs[job['companyShortName'] + job['positionName']] = job
    return newJobs.values()


def drawDistrict():
    # 按区域统计
    # 先按公司将相同职位合并
    jobs = getJobs()

    districts = {}
    labels = []
    for data in jobs:
        districts[data['district']] = districts.get(data['district'], 0) + 1

    for key, value in districts.items():
        labels.append(key + "(" + str(value) + ")")
    sizes = districts.values()

    # 设置分离的距离，0表示不分离
    explode = []

    max_value = 0
    max_index = 0

    count = 0
    for key, value in districts.items():
        if value > max_value:
            max_value = value
            max_index = count
        count += 1

    for i in range(len(districts)):
        if i == max_index:
            # 把最大的凸显出来
            explode.append(0.1)
        else:
            explode.append(0)

    plt.pie(sizes, explode=explode, labels=labels,  autopct='%1.1f%%', shadow=True, startangle=90)

    # Equal aspect ratio 保证画出的图是正圆形
    plt.axis('equal')
    plt.title('职位区域分布情况')
    plt.show()

def drawYearSalary():
    jobs = getJobs()

    labels = []
    salary_data = []
    for job in jobs:
        # 把范围处理一下在统计
        ss = job['salary'].split('-')
        min = int(ss[0].replace('k', '')) * 1000
        max = int(ss[1].replace('k', '')) * 1000
        salary_data.append({'min': min, 'max': max, 'year': job['workYear']})
        # salarys[data['salary']] = salarys.get(data['salary'], 0) + 1


    #{"30k以上":{"3-5":0}}
    year_salary = {}

    for data in salary_data:
        for key, value in salary_range.items():
            if value['min'] <= data['min'] <= value['max'] or value['min'] <= data['max'] <= value['max']:
                year_salary.setdefault(key, {data['year']: 0})
                year_salary[key].setdefault(data['year'], 0)
                year_salary[key][data['year']] = year_salary[key][data['year']] + 1

    for key in salary_range.keys():
        labels.append(key)

    sizes = []
    for v in salary_range.values():
        sizes.append(v['count'])

    workYear = {'在校/应届': 'red', '1-3年': 'orange', '3-5年': 'blue', '5-10年': 'green', '不限': 'purple'}

    # 数据对齐
    for k, v in year_salary.items():
        for year in workYear:
            v.setdefault(year, 0)

    bottom = True
    count = np.array(sizes)
    for year, color in workYear.items():
        # 从年限维度拆分数据
        data = []
        # 数据堆叠
        for k in labels:
            data.append(year_salary[k][year])

        if bottom:
            plt.bar(labels, data, color=color, alpha=0.6, label=year)
            bottom = False
        else:
            plt.bar(labels, data, bottom=count, alpha=0.6, color=color, label=year)

        count += np.array(data)

    plt.legend()
    plt.title('薪资的工作年限分布')
    plt.show()

def drawEduSalary():
    jobs = getJobs()

    labels = []
    salary_data = []
    for job in jobs:
        # 把范围处理一下在统计
        ss = job['salary'].split('-')
        min = int(ss[0].replace('k', '')) * 1000
        max = int(ss[1].replace('k', '')) * 1000
        salary_data.append({'min': min, 'max': max, 'edu': job['education']})


    #{"30k以上":{"本科":0}}
    edu_salary = {}

    for data in salary_data:
        for key, value in salary_range.items():
            if value['min'] <= data['min'] <= value['max'] or value['min'] <= data['max'] <= value['max']:
                edu_salary.setdefault(key, {data['edu']: 0})
                edu_salary[key].setdefault(data['edu'], 0)
                edu_salary[key][data['edu']] = edu_salary[key][data['edu']] + 1

    for key in salary_range.keys():
        labels.append(key)

    sizes = []
    for v in salary_range.values():
        sizes.append(v['count'])

    edus = {'大专': 'red', '本科': 'orange', '不限': 'purple'}

    # 数据对齐
    for k, v in edu_salary.items():
        for edu in edus:
            v.setdefault(edu, 0)

    bottom = True
    count = np.array(sizes)
    for edu, color in edus.items():
        # 从年限维度拆分数据
        data = []
        # 数据堆叠
        for k in labels:
            data.append(edu_salary[k][edu])

        if bottom:
            plt.bar(labels, data, color=color, alpha=0.6, label=edu)
            bottom = False
        else:
            plt.bar(labels, data, bottom=count, alpha=0.6, color=color, label=edu)

        count += np.array(data)

    plt.legend()
    plt.title('薪资的学历分布')
    plt.show()

def drawDistrictSalary():
    jobs = getJobs()

    labels = []
    salary_data = []
    for job in jobs:
        # 把范围处理一下在统计
        ss = job['salary'].split('-')
        min = int(ss[0].replace('k', '')) * 1000
        max = int(ss[1].replace('k', '')) * 1000
        salary_data.append({'min': min, 'max': max, 'district': job['district']})


    #{"30k以上":{"本科":0}}
    district_salary = {}

    for data in salary_data:
        for key, value in salary_range.items():
            if value['min'] <= data['min'] <= value['max'] or value['min'] <= data['max'] <= value['max']:
                district_salary.setdefault(key, {data['district']: 0})
                district_salary[key].setdefault(data['district'], 0)
                district_salary[key][data['district']] = district_salary[key][data['district']] + 1

    for key in salary_range.keys():
        labels.append(key)

    sizes = []
    for v in salary_range.values():
        sizes.append(v['count'])

    d = []
    for job in jobs:
        d.append(job['district'])

    # {'西湖区', '淳安县', '下沙', '滨江区', '余杭区', '江干区', '桐庐县', '拱墅区', '萧山区'}
    d = set(d)

    districts = {'西湖区': 'red', '滨江区': 'orange', '下沙': 'pink', '余杭区': 'grey', '江干区': 'brown', '拱墅区': 'black', '萧山区': 'green', '桐庐县': 'purple', '淳安县': 'yellow'}

    # 数据对齐
    for k, v in district_salary.items():
        for edu in districts:
            v.setdefault(edu, 0)

    bottom = True
    count = np.array(sizes)
    for edu, color in districts.items():
        # 从年限维度拆分数据
        data = []
        # 数据堆叠
        for k in labels:
            data.append(district_salary[k][edu])

        if bottom:
            plt.bar(labels, data, color=color, alpha=0.6, label=edu)
            bottom = False
        else:
            plt.bar(labels, data, bottom=count, alpha=0.6, color=color, label=edu)

        count += np.array(data)

    plt.legend()
    plt.title('薪资的地区分布')
    plt.show()


# url = "https://www.lagou.com/jobs/list_%E4%BC%9A%E8%AE%A1/p-city_6?&cl=false&fromSearch=true&labelWords=&suginput="
# url = "https://www.lagou.com/jobs/8237213.html?show=d9536b2de75d43c69abb6e1e0630a8e5"
# # 开启chromedriver
# browser = webdriver.Chrome()
# browser.get(url)
# browser.implicitly_wait(10)
#
# # 写入源html
# f = open('lagou.txt', 'w', encoding='utf8')
# raw_html = browser.page_source
# f.write(raw_html)
# f.close()
#


def get_json(url, num):
    # fixme (1) '''''从网页获取JSON,使用POST请求,加上头部信息'''
    # my_headers = {
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
    #     'Host': 'www.lagou.com',
    #     'Referer': 'https://www.lagou.com/jobs/list_%E6%95%B0%E6%8D%AE%E5%88%86%E6%9E%90?labelWords=&fromSearch=true&suginput=',
    #     'X-Anit-Forge-Code': '0',
    #     'X-Anit-Forge-Token': 'None',
    #     'X-Requested-With': 'XMLHttpRequest',
    #     'cookie': 'user_trace_token=20210318192205-97eed590-9e0b-41ce-aaf3-e05b4a5c0dbd; JSESSIONID=ABAAABAABAGABFA0096BD51F33F05CF510EF7781D45CEE3; WEBTJ-ID=20210318%E4%B8%8B%E5%8D%887:22:08192208-17845124659513-040c4d5212b6d7-33697709-1296000-1784512465a5ad; _ga=GA1.2.389223511.1616066530; LGUID=20210318192210-229c0ce9-0b71-4b67-a299-c81ff13f371b; sensorsdata2015session=%7B%7D; RECOMMEND_TIP=true; index_location_city=%E6%9D%AD%E5%B7%9E; _gid=GA1.2.152860951.1616328044; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1616066530,1616328052; __lg_stoken__=b9bde51040dc9ec29072fdb9254c20dd6d36775e4973de30bd6fd6c8e077c9850a6b57556c029c623eb98a4ecb11ffd8078268e8a894fab6fa0798c838504237de78944fb21b; X_MIDDLE_TOKEN=3c1d2fb2e97ef2aa954099dedfedbbad; TG-TRACK-CODE=search_code; gate_login_token=521d39a1f3e7b08c3a46e0b5d024f6dc079e10082b46ffc9668f1d6166f37215; _putrc=1D0A9BCDB877FC11123F89F2B170EADC; login=true; unick=%E7%94%A8%E6%88%B79536; hasDeliver=0; privacyPolicyPopup=false; showExpriedIndex=1; showExpriedCompanyHome=1; showExpriedMyPublish=1; LGSID=20210322011021-22940cae-3a24-4608-bab5-afc1bbc609d7; PRE_UTM=; PRE_HOST=; PRE_SITE=https%3A%2F%2Fwww.lagou.com; PRE_LAND=https%3A%2F%2Fwww.lagou.com%2Fjobs%2Flist%5F%25E4%25BC%259A%25E8%25AE%25A1%2Fp-city%5F6%3F%26cl%3Dfalse%26fromSearch%3Dtrue%26labelWords%3D%26suginput%3D; _gat=1; X_HTTP_TOKEN=6a724cad9c93210a5266436161014c0f701d89af3d; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2221042125%22%2C%22first_id%22%3A%2217845124c622b-02927cfa9ad709-33697709-1296000-17845124c633db%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%2C%22%24os%22%3A%22MacOS%22%2C%22%24browser%22%3A%22Chrome%22%2C%22%24browser_version%22%3A%2288.0.4324.192%22%2C%22lagou_company_id%22%3A%22%22%7D%2C%22%24device_id%22%3A%2217845124c622b-02927cfa9ad709-33697709-1296000-17845124c633db%22%7D; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1616346627; SEARCH_ID=ddcb7f53fa224cf58d43a0ea1df2d67e; LGRID=20210322011037-1c640d6a-8f35-40d3-8498-2dbe5a855d09'
    # }
    #
    # my_data = {
    #     'first': 'true',
    #     'pn': num,
    #     'kd': '会计'}
    #
    #
    # res = requests.post(url, headers=my_headers, data=my_data)
    # res.raise_for_status()
    # res.encoding = 'utf-8'
    # # 得到包含职位信息的字典
    # page = res.json()

    # fixme (2) 切换到从文件中加载职位信息
    f = open("original_jobs_data.txt")
    s = f.read()
    f.close()
    page = json.loads(s)

    return page


class Data:
    jobs = []
    position = {}


def buildMapData():
    companies = {}
    original_position_list = []

    # for i in range(8):
    for i in range(1):
        page = get_json(
            "https://www.lagou.com/jobs/positionAjax.json?city=%E6%9D%AD%E5%B7%9E&needAddtionalResult=false",
            i + 1)
        # fixme (1)
        # positions = page['content']['positionResult']['result']

        # fixme (2)
        positions = page

        for p in positions:
            original_position_list.append(p)

            # {content:"我的备注test",title:"test",imageOffset: {width:0,height:3},position:{lat:30.309139,lng:120.125877}}
            name = p['companyShortName']
            data = companies.get(name)
            if data == None:
                data = Data()
                data.jobs = []
                data.position = {'lat': p['latitude'], 'lng': p['longitude']}
                companies[name] = data

            data.jobs.append("职位:" + p['positionName'] + ", 薪资：" + p['salary'] + ", 年限：" + p['workYear'])

        print("process page:" + str(i + 1) + " is OK!")
        # fixme (1) 打开注释, 暂停10s
        # fixme (2) 读取本地文件, 无需暂停
        # time.sleep(10)

    # fixme (1) 需要打开注释: 把原始数据写入保存, 避免多次爬数据被屏蔽
    # fixme (2) 注释掉文件写入
    # f = open("original_jobs_data.txt", "w")
    # f.write(json.dumps(original_position_list))
    # f.close()

    rows = []
    for key, data in companies.items():
        row = {
            'title': key,
            "content": '<br>'.join(set(data.jobs)),
            'imageOffset': {'width': 0, 'height': 3},
            'position': {'lat': data.position['lat'], 'lng': data.position['lng']}
        }
        rows.append(row)

    # print(json.dumps(rows))

    f = open("jobs.txt", 'w')
    f.write(json.dumps(rows))
    f.close()

    print("write file is ok")


if __name__ == "__main__":
    # createExcel()
    drawDistrict()
    drawYearSalary()
    drawEduSalary()
    drawDistrictSalary()
    print("任务完成")

# header = {
#     "Origin": "http://www.lagou.com",
#     "Accept": "application/json, text/javascript, */*; q=0.01",
#     "Referer": "http://www.lagou.com",
#     'User-Agent': "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.192 Safari/537.36",
#     'Cookie': "user_trace_token=20210318192205-97eed590-9e0b-41ce-aaf3-e05b4a5c0dbd; JSESSIONID=ABAAABAABAGABFA0096BD51F33F05CF510EF7781D45CEE3; WEBTJ-ID=20210318%E4%B8%8B%E5%8D%887:22:08192208-17845124659513-040c4d5212b6d7-33697709-1296000-1784512465a5ad; _ga=GA1.2.389223511.1616066530; LGUID=20210318192210-229c0ce9-0b71-4b67-a299-c81ff13f371b; sensorsdata2015session=%7B%7D; RECOMMEND_TIP=true; index_location_city=%E6%9D%AD%E5%B7%9E; _gid=GA1.2.152860951.1616328044; LGSID=20210321200043-688cd7b0-2399-4268-b211-c6f11a8aaac5; Hm_lvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1616066530,1616328052; __lg_stoken__=b9bde51040dc9ec29072fdb9254c20dd6d36775e4973de30bd6fd6c8e077c9850a6b57556c029c623eb98a4ecb11ffd8078268e8a894fab6fa0798c838504237de78944fb21b; X_MIDDLE_TOKEN=3c1d2fb2e97ef2aa954099dedfedbbad; TG-TRACK-CODE=search_code; gate_login_token=521d39a1f3e7b08c3a46e0b5d024f6dc079e10082b46ffc9668f1d6166f37215; _putrc=1D0A9BCDB877FC11123F89F2B170EADC; login=true; unick=%E7%94%A8%E6%88%B79536; hasDeliver=0; privacyPolicyPopup=false; X_HTTP_TOKEN=6a724cad9c93210a5699236161014c0f701d89af3d; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2221042125%22%2C%22first_id%22%3A%2217845124c622b-02927cfa9ad709-33697709-1296000-17845124c633db%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%2C%22%24os%22%3A%22MacOS%22%2C%22%24browser%22%3A%22Chrome%22%2C%22%24browser_version%22%3A%2288.0.4324.192%22%2C%22lagou_company_id%22%3A%22%22%7D%2C%22%24device_id%22%3A%2217845124c622b-02927cfa9ad709-33697709-1296000-17845124c633db%22%7D; Hm_lpvt_4233e74dff0ae5bd0a3d81c6ccf756e6=1616329967; SEARCH_ID=b0f80c1ad5ec4ecab59ebd7c8f30498f; LGRID=20210321203443-697607d6-3ee7-4e22-b0d0-838dc3ec5b57"
# }
# url = 'https://www.lagou.com/jobs/positionAjax.json?city=杭州&needAddtionalResult=false'
# r = requests.post(url=url, data={'pn': '1', 'kd': '会计'}, headers=header)
# print(r.text)


# # 通过requests获取数据
# r = requests.get('http://echarts.apache.org/examples/data/asset/data/hangzhou-tracks.json')
# data = r.json()
#
# data_pair = []
#
# # 新建一个BMap对象
# bmap = BMap()
#
# for i, item in enumerate([j for i in data for j in i ]):
#     # 新增坐标点
#     bmap.add_coordinate(i, item['coord'][0], item['coord'][1])
#     data_pair.append((i, 1))
#
# bmap.add_schema(
#     # 需要申请一个AK
#     baidu_ak='293LxGqPbsciHuzbnSmBzljnItUpCAH5',
#     # 地图缩放比例
#     zoom=14,
#     # 显示地图中心坐标点
#     center=[120.13066322374, 30.240018034923])
#
# # 添加数据
# bmap.add("门店数", data_pair,
#          type_='heatmap')
#
# # 数据标签不显示
# bmap.set_series_opts(label_opts=opts.LabelOpts(is_show=False))
# bmap.set_global_opts(
#     visualmap_opts=opts.VisualMapOpts(min_=0, max_=50,
#                                       # 颜色效果借用Echarts示例效果
#                                       range_color=['blue', 'blue', 'green', 'yellow', 'red']),
#     # 图例不显示
#     legend_opts=opts.LegendOpts(is_show=False),
#     title_opts=opts.TitleOpts(title="杭州热门步行路线"))
#
# bmap.render()
